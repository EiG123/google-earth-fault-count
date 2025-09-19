import logging
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill


def write_results_to_excel(points_df, redline_summary, threshold_m, output_path=None, use_detail_count=False):
    """
    เขียนผลไปเป็น Excel:
      - sheet 'points_summary' = สรุปเส้น + นับแยกรายเดือน พร้อม hyperlink
      - sheet per redline = รายละเอียด (ticket, ref, distance_m/อื่นๆ)
      - sheet 'statistics' = ข้อมูลนับซ้ำและ Duplicate Rate
      - ชื่อไฟล์จะใส่ threshold_m และวันที่เวลาปัจจุบัน
    
    Args:
        use_detail_count (bool): ถ้า True ใช้ points_by_details, ถ้า False ใช้ points (coordinate-based)
    """
    # ตั้งชื่อไฟล์ถ้าไม่ได้ส่งมา
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        count_type = "details" if use_detail_count else "coords"
        output_path = f"results_points_redlines_{threshold_m}m_{count_type}_{timestamp}.xlsx"

    # -------------------
    # 1) Summary
    # -------------------
    summary_rows = []
    for rl_name, info in redline_summary.items():
        pts_for_summary = info["raw_matches"]
        raw_matches = info["raw_matches"]

        if not pts_for_summary:
            summary_rows.append({
                "เส้นสายไฟ": rl_name,
                "จำนวนจุดทั้งหมด": 0,
                "มกราคม": 0, "กุมภาพันธ์": 0, "มีนาคม": 0, "เมษายน": 0,
                "พฤษภาคม": 0, "มิถุนายน": 0, "กรกฎาคม": 0, "สิงหาคม": 0,
                "กันยายน": 0, "ตุลาคม": 0, "พฤศจิกายน": 0, "ธันวาคม": 0,
                "อื่นๆ": 0,
                "ระยะเฉลี่ย (m)": 0
            })
            continue

        df_summary = pd.DataFrame(pts_for_summary)
        df_raw = pd.DataFrame(raw_matches) if raw_matches else pd.DataFrame()

        # ตรวจสอบว่ามีคอลัมน์ key หรือไม่
        if "key" in df_summary.columns:
            jan = (df_summary["key"] == "มกราคม").sum()
            feb = (df_summary["key"] == "กุมภาพันธ์").sum()
            mar = (df_summary["key"] == "มีนาคม").sum()
            apr = (df_summary["key"] == "เมษายน").sum()
            may = (df_summary["key"] == "พฤษภาคม").sum()
            jun = (df_summary["key"] == "มิถุนายน").sum()
            jul = (df_summary["key"] == "กรกฎาคม").sum()
            aug = (df_summary["key"] == "สิงหาคม").sum()
            sep = (df_summary["key"] == "กันยายน").sum()
            octb = (df_summary["key"] == "ตุลาคม").sum()
            nov = (df_summary["key"] == "พฤศจิกายน").sum()
            dec = (df_summary["key"] == "ธันวาคม").sum()

            # นับ "อื่นๆ"
            months = {"มกราคม","กุมภาพันธ์","มีนาคม","เมษายน","พฤษภาคม","มิถุนายน",
                      "กรกฎาคม","สิงหาคม","กันยายน","ตุลาคม","พฤศจิกายน","ธันวาคม"}
            other_count = df_summary[~df_summary["key"].isin(months)].shape[0]
        else:
            # ถ้าไม่มี group → ตั้งค่าเป็นศูนย์หมด
            jan = feb = mar = apr = may = jun = jul = aug = sep = octb = nov = dec = 0
            other_count = 0


        # คำนวณระยะเฉลี่ย
        avg_distance = df_raw["distance_m"].mean() if "distance_m" in df_raw.columns and len(df_raw) > 0 else 0

        summary_rows.append({
            "เส้นสายไฟ": rl_name,
            "จำนวนจุดทั้งหมด": len(df_summary),
            "มกราคม": jan, "กุมภาพันธ์": feb, "มีนาคม": mar, "เมษายน": apr,
            "พฤษภาคม": may, "มิถุนายน": jun, "กรกฎาคม": jul, "สิงหาคม": aug,
            "กันยายน": sep, "ตุลาคม": octb, "พฤศจิกายน": nov, "ธันวาคม": dec,
            "อื่นๆ": other_count,
            "ระยะเฉลี่ย (m)": round(avg_distance, 2)
        })

    summary_df = pd.DataFrame(summary_rows)

    # แถวรวม
    total_row = {
        "เส้นสายไฟ": "รวมทั้งหมด",
        "จำนวนจุดทั้งหมด": summary_df["จำนวนจุดทั้งหมด"].sum(),
        "มกราคม": summary_df["มกราคม"].sum(),
        "กุมภาพันธ์": summary_df["กุมภาพันธ์"].sum(),
        "มีนาคม": summary_df["มีนาคม"].sum(),
        "เมษายน": summary_df["เมษายน"].sum(),
        "พฤษภาคม": summary_df["พฤษภาคม"].sum(),
        "มิถุนายน": summary_df["มิถุนายน"].sum(),
        "กรกฎาคม": summary_df["กรกฎาคม"].sum(),
        "สิงหาคม": summary_df["สิงหาคม"].sum(),
        "กันยายน": summary_df["กันยายน"].sum(),
        "ตุลาคม": summary_df["ตุลาคม"].sum(),
        "พฤศจิกายน": summary_df["พฤศจิกายน"].sum(),
        "ธันวาคม": summary_df["ธันวาคม"].sum(),
        "อื่นๆ": summary_df["อื่นๆ"].sum(),
        "ระยะเฉลี่ย (m)": round(summary_df["ระยะเฉลี่ย (m)"].mean(), 2)
    }
    summary_df = pd.concat([summary_df, pd.DataFrame([total_row])], ignore_index=True)

    # -------------------
    # 2) เขียนลง Excel
    # -------------------
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # summary
        summary_df.to_excel(writer, sheet_name="points_summary", index=False)

        # per redline
        for rl_name, info in redline_summary.items():
            all_matches = info["raw_matches"]
            if not all_matches:
                continue

            df = pd.DataFrame(all_matches)
            if "distance_m" in df.columns:
                df = df.sort_values("distance_m")

            safe_name = rl_name.replace("/", "_").replace("\\", "_").replace(":", "_")
            sheet_name = (safe_name[:28] + "...") if len(safe_name) > 31 else safe_name

            existing_sheets = writer.book.sheetnames
            if sheet_name in existing_sheets:
                suffix = 1
                base_name = safe_name[:25] if len(safe_name) > 25 else safe_name
                while sheet_name in existing_sheets:
                    sheet_name = f"{base_name}_{suffix}"
                    suffix += 1

            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # statistics
        stats_data = []
        for rl_name, info in redline_summary.items():
            stats_data.append({
                "เส้นสายไฟ": rl_name,
                "Count by Coordinates": info["count_by_coords"],
                "Count by Details": info["count_by_details"],
                "Total Matches": info["total_matches"],
                "Duplicate Rate (%)": round(
                    ((info["total_matches"] - info["count_by_coords"]) / info["total_matches"] * 100)
                    if info["total_matches"] > 0 else 0, 2
                )
            })
        stats_df = pd.DataFrame(stats_data)
        stats_df.to_excel(writer, sheet_name="statistics", index=False)

    # -------------------
    # 3) ปรับแต่งด้วย openpyxl
    # -------------------
    wb = load_workbook(output_path)
    ws_summary = wb["points_summary"]

    # เพิ่ม hyperlink
    for row_idx in range(2, len(summary_df) + 1):  # รวม header = 1
        rl_name = ws_summary.cell(row=row_idx, column=1).value
        if rl_name == "รวมทั้งหมด":
            continue

        safe_name = rl_name.replace("/", "_").replace("\\", "_").replace(":", "_")
        target_sheet = None
        if safe_name in wb.sheetnames:
            target_sheet = safe_name
        else:
            for sname in wb.sheetnames:
                if sname.startswith(safe_name[:25]):
                    target_sheet = sname
                    break

        if target_sheet:
            cell = ws_summary.cell(row=row_idx, column=1)
            cell.hyperlink = f"#'{target_sheet}'!A1"
            cell.style = "Hyperlink"

    # จัด bold + fill แถวรวม
    total_row_idx = len(summary_df) + 1
    for col in range(1, ws_summary.max_column + 1):
        cell = ws_summary.cell(row=total_row_idx, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # ปรับความกว้าง column ทุก sheet
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col)
            for cell in ws[col_letter]:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(output_path)

    # -------------------
    # 4) Logging
    # -------------------
    total_points_coords = sum(info["count_by_coords"] for info in redline_summary.values())
    total_points_details = sum(info["count_by_details"] for info in redline_summary.values())
    total_matches = sum(info["total_matches"] for info in redline_summary.values())

    logging.info("บันทึกผลเป็น Excel ที่: %s", output_path)
    logging.info("สถิติ: จุดที่ไม่ซ้ำ (coords)=%d, จุดที่ไม่ซ้ำ (details)=%d, matches รวม=%d",
                 total_points_coords, total_points_details, total_matches)

    return output_path
